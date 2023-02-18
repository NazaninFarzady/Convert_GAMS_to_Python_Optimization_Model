'''
This code script will do the following:
1. Read the parameters' value from 'Setting.txt' files
2. Set the working directory the data set folder
3. Read the Excel files,
4. Read the defined_names, the range of data values in Excel files from different Tabs
5. Store each defined_name as dataframe and save as 'CSV' format in new sub-folder
6. Recall dataframes and clean considering their specific structure
7. Change data types in dataframe according to 'Data Documentation' and replace 0 values
8. Compute the parameters' value according to the GAMS formulations/codes
9. Create dataframes for mapping and assignments of Sets
10.Create dataframe for each parameter with its related indices
11.Save new dataframe as 'CSV' format in new sub_folder for Pyomo model
'''
# Import libraries
import sys
import os
import glob
import csv
from os import listdir
from os.path import splitext, basename,isfile, join
import pandas as pd
import openpyxl
import numpy as np
import itertools
from openpyxl import load_workbook,workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from functools import reduce


# store the 'Setting.txt' file as a variable
file = 'Setting.txt'

# Function to read, clean and store parameters' value with their related indices

try:
    def clean_data(file):
        ###################### Read the setting text file which includes Global parameters #########################
        param = {}  # an empty dictionary to store parameters value from text file
        try:
            current_path = os.getcwd()  # get the current path
            with open(file) as f:  # open text file as 'f'
                for line in f:  # read each line in file
                    splitLine = line.split(':')  # in each line, parameter's name and value are separated by (:)
                    param[splitLine[0]] = ",".join(splitLine[1:]).strip('\n')  # Get the first term as key and the second as value, then it stores the value in a variable and remove the extra character (\n)
            for key, val in param.items():  # get the parameters' value and save in variables for further reference
                if key == 'Main':  # The main data folder
                    main = val
                if key == 'Data':  # The dataset sub_folder
                    data_test = val
                if key == 'Scenario_WEO':  # The scenario of 'WEO'
                    weo = val
                if key == 'Scenario_SETNav':  # The scenario of 'SETNav'
                    setnav = val
                if key == 'Year':  # The value of time period
                    last_year = val
        except FileNotFoundError:
            print("Directory: {0} does not exist".format(file))  # Print this message if file is nto found
        # print(type(last_year))
        ####################### Read the excel files from data test folder directory #########################
        try:
            data = main  # get the main folder of all data ('data'), it reads from setting text file
            os.chdir(data)  # change directory the 'data' main folder
            data_path = os.getcwd()  # get directory after changed to 'data' folder
            sub_dir = data_test  # get the sub_folder directory (test folders), reads from setting text file
            test_dir = os.path.join(data_path, sub_dir)  # join test sub_folder to the changed directory
            temp_dir = os.path.join(test_dir,
                                    'temp')  # create new sub_folder to store defined_ranges as dataframe in 'CSV' format
            if not os.path.exists(temp_dir):  # If the name of sub_folder is not in desired directory, create it
                os.makedirs(os.path.join(test_dir, temp_dir))

            files = os.listdir(test_dir)  # list the files inside test folder
            files_xlsx = [f for f in files if f[-4:] == "xlsx"]  # Select only 'xlsx' files as a list
            list_wb = []  # empty list to store workbooks
            for excel in files_xlsx:  # for each excel file in the list of files,
                excel_dir = test_dir + '\\' + excel  # set a path of excel file
                wb = load_workbook(filename=excel_dir, data_only=True)  # ead the workbook from set path
                list_wb.append(wb)  # add the workbook to the empty list

            ############################ Read  defined name ranges from excel file ##########################
            results = []  # empty list to store the defined names
            for workbook in list_wb:  # for each workbook from the list
                for item in workbook.defined_names.definedName:  # get each item of defined range with its attributes
                    sh = workbook.sheetnames  # create sheet name variable for later checking of ranges
                    if '!#REF!' not in item.attr_text:  # check for reference cell values which are not a range and has specific character
                        for title, coords in item.destinations:  # get the name and range attribute
                            if title in sh:  # check if the title is in sheet range of workbook
                                sheet = workbook[title]  # Set the sheet as workbook title attribute
                                range = sheet[
                                    coords]  # set the range value as the range of values in the sheet of workbook, the attribute of coords
                                if isinstance(range, tuple):  # check if the range is a single value or a range value
                                    for row in range:  # for row in range of values
                                        name = item.name  # each range's name is the name of defined_range
                                        item_rgn = ([str(cell.value) for cell in
                                                     row])  # the range of value if the value of the rows
                                        results.append(
                                            (name, item_rgn))  # add the name and the range of value ro the list
                                else:  # if the range is single value
                                    results.append((item.name,
                                                    range.value))  # get the name of range and the single value as a tuple then add to the list
                            else:
                                print(
                                    'The Range is loaded from different workbook')  # if the defined_name in the excel is loaded from another workbooks, print this message
                    else:
                        # print('Range is not defined and contained character (!#REF!)')
                        pass

                    # Group the range names and their values
                    groups = {}  # empty dictionary to store ranges' name with their values (data values in each defined range)
                    for group, value in results:  # for each range's name and its value as a group,
                        if group not in groups:  # if the group is not in the list
                            groups.update({group: [value]})  # add the group as key and value as value of dictionary
                        else:
                            groups[group].append(value)  # do the same as above for pair items in groups

                    # save key-value pairs of dictionary into dataframe as 'csv' file format
                    for key, val in groups.items():
                        df = pd.DataFrame(groups[key])  # make dataframe from groups (dictionary) based on its keys
                        df.to_csv(f'{test_dir}/temp/{key}.csv', index=False)  # header=0)
                        # save each dataframe with name of dictionary key in 'csv' format without header
                        # save defined_ranges of data as 'csv' files in sub_folder (the output of this part of the code)
            # print('Data folder directory:{0}'.format(temp_dir))

        except FileNotFoundError:
            print("Directory: {0} does not exist".format(
                data_path))  # Print this message if the file or directory could not be be found

        ################## Read the range of data values from 'csv' files as dataframe ###############
        try:
            files = glob.glob(
                temp_dir + '/*.csv')  # get 'csv' files from temp sub_folder (these are the stored files from previous code block, as defined_ranges
            dfs = []  # create an empty list to store dataframes
            for file in dfs:  # for each file is the list
                df = pd.read_csv(file)  # read the csv file
                dfs.append(df)  # add the dataframe to the list

            # For each element in dfs, will split the name of the file, and make dfs as a dictionary with the (files' name) as key and dataframes as value
            dfs = {splitext(basename(df))[0]: pd.read_csv(df) for df in files}

            # Then we can call each dataframe based on the defined_name in Excel. For example 'PriceCalib' is structured as dataframe containing the data value from excel file
            for k, v in dfs.items():
                globals()[k] = v
            # print(Arcs)
        except FileNotFoundError:
            print("Directory: {0} does not exist".format(
                files))  # print this message if the file or folder could not be found

            ################## Read data values from dataframes ###############
        '''## Since the defined_name ranges do not have same structure, I was not able to define a general function to modify them at once
            ## Therefore, each dataframe i.e., defined_names are modified separately as it needed
            ## Consequently, it is better that the name of defined_name in Name manager of Excel file, be same as the ones that code is written based on.
            ## Otherwise, the name should be change from here to the rest of the script
            ## To clean dataframes which the first row is labeled with numbers after reading the 'csv' file, the header is placed in the second row.
            ## The following function cleans this kind of dataframe, gets the header from the second row, and set as columns names/dataframe header
            '''

        def clean_header(df):
            new_header = df.iloc[0]
            df = df[1:]
            df.columns = new_header
            df = df.reset_index(drop=True)
            return df

        '''# To apply this function it calls dataframe (the defined_name range) like this --> clean_header(Arcs)
            # Note: This function is not work with all dataframe, because in excel files ranges are defined differently.
            # Some of the ranges does not cover the header of columns.
            '''

        #################  Reading Nodes Data ########################
        '''##  Dataframe 'Nodes' contains the name of all nodes as well as dataframe 'Arcs' which have the name of all arcs.
            ##  These data values would be read without changing by the pyomo model as set of nodes and arcs.
            ##  Therefore, these two dataframes will be save as they are in new sub_folder for pyomo model

            ## 'NodesData' contains the data values for the name of nodes, the types, country's name and regions.
            ##  This data is also needed to map nodes with region and country that they are related each other.
            '''
        ##  First clean the dataframe header
        try:
            NodesData_cleaned = clean_header(NodesData)

            # Then determine the nodes type
            # Because of some overlapping to avoid complication, one simple way to determine nodes type could be as follows:
            # Saving the list of Node types which later will be read by the pyomo model as set for indexing parameters and variables

            for row in NodesData_cleaned:
                n_cons = NodesData_cleaned.loc[NodesData_cleaned[
                                                   'C'] == '1', 'N']  # List of Consumption Nodes, (nodes with consumption columns values (C) == 1)
                n_prod = NodesData_cleaned.loc[NodesData_cleaned[
                                                   'P'] == '1', 'N']  # List of Production Nodes, (nodes with production columns (P) == 1)
                n_liq = NodesData_cleaned.loc[
                    NodesData_cleaned['L'] == '1', 'N']  # List of liquefaction Nodes, (nodes with L columns (L) == 1)
                n_reg = NodesData_cleaned.loc[
                    NodesData_cleaned['R'] == '1', 'N']  # List of regasifaction Nodes, (nodes with R columns (R) == 1)
                countries = NodesData_cleaned[
                    'CN'].drop_duplicates()  # is the list of countries, but it might not be needed as pyomo Set() component
                suppliers = NodesData_cleaned.loc[
                    NodesData_cleaned['P'] == '1', 'CN']  # List of supplier Nodes, (nodes with production columns == 1)

            ### Map and select corresponding Nodes with the countries and Regions

            map_n_cn = NodesData_cleaned[['N', 'CN']]  # dataframe of countries with their nodes
            map_cn_rgn = NodesData_cleaned[['CN', 'Rgn']]  # Regions with the countries inside
            map_n_rgn = NodesData_cleaned[['N', 'Rgn']]  # dataframe of regions and the nodes inside them
            map_n_cn_rgn = NodesData_cleaned[
                ['N', 'CN', 'Rgn']]  # dataframe of regions with the countries inside them with their nodes
        except Exception as e:
            print(e)
            print('Check Reading Nodes Data block')
            # We are done with NodesData table from Excel file
        ##############################################################################################
        ################################### TIME PERIOD DATA VALE ####################################
        # Reading data values for time period from 'Years' dataframe
        try:
            y_df = Years.copy()  # make a copy from main Years dataframe
            y_df = y_df.rename(columns={'0': 'Year'}).reset_index(
                drop=True)  # change the column name from '0' to 'Year'
            y_df['Year'] = y_df['Year'].astype(
                str)  # convert value under the column to str obj suits value for the index set

            ##### Function to get the position of the year period ####################
            ## NOTE: In python indexing starts from 0 then this index position is equivalent of the terms in GAMS (ORD(y)-1)
            ## Therefore, in equation this term (ORD(y)-1) is replaced by the get_indx(y)
            # This function gets Years, as dataframe, and the value of last_year parameter which is read from Setting test file, and return the position of the year
            def get_indx(df, y):
                indx_y = y_df[y_df['Year'] == last_year].index.values.astype(int)[0]
                return indx_y

                # get the index of 'year'

            y_index = get_indx(y_df, last_year)
        except IndexError as ie:
            print(ie)
            print('Check the TIME PERIOD DATA VALE block for year')
        ############ Production Resource data ###################
        # Reading data values for the Production Resource from 'Resources' dataframe
        try:
            df_r = Resources.copy()
            df_r = df_r.rename(columns={'0': 'R'}).reset_index(drop=True)  # Rename the header from '0' to 'R'
            df_r['R'] = df_r['R'].astype(str)  # convert data type to an object (string)
        except Exception as e:
            print(e)
            print('Check the Production Resource data block')

        ################# OTHER PARAMETERS' VALUE ###############
        # Read the values for scalar parameters from 'Other' dataframe
        try:
            Others = Other.rename(
                columns={'0': 'Param', '1': 'Value'})  # rename the headers, '0' to 'Param' and '1' to 'Value'

            # Replace 'EPS' in column of Parameter values ('Value') with zero, convert data type to (int)
            Others['Value'] = Others['Value'].replace('EPS', 0).astype(float)

            # Convert dataframe to dictionary with (Key=parameter, Value=value)
            # Remove the index from dataframe, transpose and convert to dictionary as list format
            Others = Others.set_index('Param').T.to_dict('series')

            # Get the value for scalar parameters the step year, price inflator, cost inflator, Pipeline offshore multiplication, and Discount rate
            step_y = int(Others['YearStep'])
            pr_infl = float(Others['PriceInfl'])
            cost_infl = float(Others['CostInfl'])
            PipeOffshMult = float(Others['PipeOffshMult'])
            disc_rate = float(Others['DiscRate'])
        except Exception as e:
            print(e)
            print("Check the OTHER PARAMETERS' VALUE data")

        ##################### in_arcs file in GAMS #####################
        ######################### Reading ARCS DATA  #########################
        # Using function to clean the header
        try:
            ArcsData_cleaned = clean_header(ArcsData)
            ArcsData_cleaned.columns = ArcsData_cleaned.columns.astype(
                str)  # change the column name to string type to make sure all are in same type
            ArcsData_cleaned.loc[:, 'len':] = ArcsData_cleaned.loc[:, 'len':].replace({'EPS': 0, 'None': 0}).astype(
                float)  # slice of dataframe from 'len' to the end, replace 'EPS' with 0 and change data type tp 'float'

            '''# The data frame 'ArcsData' includes the other information about the Arcs except vessels.
            # To get the name of Vessels, dataframe 'Arcs' used and to get related nodes and other information other dataframe (LNGDistances) is used
            # The LNG data is cleaned in later section
            # The following function will return the type of Pipe, Liq or Regas according the value of columns 'P','L' and 'R'.
            # By applying this function to dataframe, it creates additional column with Arcs type
            '''

            def alert(ArcsData_cleaned):
                if ArcsData_cleaned['P'] == '1':
                    return 'Pipe'
                elif ArcsData_cleaned['L'] == '1':
                    return 'Liq'
                elif ArcsData_cleaned['R'] == '1':
                    return 'Regas'
                else:
                    return 'Undefined'

            ArcsData_cleaned["Arc_Type"] = ArcsData_cleaned.apply(alert, axis=1)

            # Map the arcs with start and end node, separately for each type of Arc
            # pass the type of arcs as 'Pipe', 'Liq', 'Regas' and function returns a dataframe include of arc's name, its start and end node
            def arc_node(Arctype):
                return ArcsData_cleaned.loc[ArcsData_cleaned['Arc_Type'] == Arctype, ['Arc', 'Start', 'End']]

            pipe_nod = arc_node('Pipe')  # dataframe of pipeline arc with start node (no) and end node (ni)
            liq_nod = arc_node('Liq')
            reg_nod = arc_node('Regas')

        except Exception as e:
            print(e)
            print('Check the Reading ARCS DATA block')
        ##### ARCS PARAMETER Computation ############
        ## This part is equivalent of 'in_arcs' GAMS file
        ## First, the scalar parameter that are related to arcs, is added to the dataframe.
        ## Then the next parameters are calculated.

        try:
            # Create an empty list for value of arcs' scalar parameters
            cost_a_base = []  # The basic usage cost for each arc type in order to Add to dataframe
            l_a_base = []  # The basic loss fraction for each arc type
            inv_a_base = []  # The basic investment cost for each arc type

            # Loop over the rows of each arc type and assign value of base cost
            for row in ArcsData_cleaned['P']:
                if row == '1':
                    cost_a_base.append(int(Others['BFPipe']))
                    l_a_base.append(float(Others['BLPipe']))
                    inv_a_base.append(int(Others['BICPipe']))

            for row in ArcsData_cleaned['L']:
                if row == '1':
                    cost_a_base.append(int(Others['BFLiq']))
                    l_a_base.append(float(Others['BLLiq']))
                    inv_a_base.append(int(Others['BICLiq']))

            for row in ArcsData_cleaned['R']:
                if row == '1':
                    cost_a_base.append(int(Others['BFReg']))
                    l_a_base.append(float(Others['BLReg']))
                    inv_a_base.append(int(Others['BICReg']))

            # Add the list of Arcs' parameters to the dataframe
            ArcsData_cleaned['cost_a_base'] = cost_a_base
            ArcsData_cleaned['cost_a_base'] = ArcsData_cleaned['cost_a_base'].replace(
                {'EPS': float(0)})  # Set EPS value to 0
            ArcsData_cleaned['l_a_base'] = l_a_base
            ArcsData_cleaned['l_a_base'] = ArcsData_cleaned['l_a_base'].replace({'EPS': float(0)})  # Set EPS value to 0
            ArcsData_cleaned['inv_a_base'] = inv_a_base
            ArcsData_cleaned['inv_a_base'] = ArcsData_cleaned['inv_a_base'].replace(
                {'EPS': float(0)})  # Set EPS value to 0

            ## Note: When the code reads the Excel file and then 'CSV' for 'ArcsData', the header of year multiplication factor for capacity, is string.
            ## To get the value of '2015' column for computing yearly capacity, the argument that will be pass to function, should be string, as '2015'
            # This function compute parameters that are affected by year
            def arc_customize(y):
                # some of these parameters are not affected by position of Year, such as loss rate
                arcs_param = ArcsData_cleaned.copy()

                # arcs_param['l_a'] = round((arcs_param['len'] * arcs_param['l_a_base']),3) #Loss Fraction of Arcs
                arcs_param['l_a'] = arcs_param['len'] * arcs_param['l_a_base']  # Loss Fraction of Arcs

                # arcs_param['cap_a'] = round((arcs_param[y] / (1 - arcs_param['l_a'])),2)  # Arc cap BCM/yr
                arcs_param['cap_a'] = arcs_param[y] / (1 - arcs_param['l_a'])  # Arc cap BCM/yr

                y_index = get_indx(y_df, y)  # Get the index position of the Year
                arcs_param['cost_a'] = arcs_param['cost_a_base'] * arcs_param['c_cal'] * (1 - arcs_param['l_a']) * (
                            arcs_param['len'] + max(0, (PipeOffshMult - 1)) * arcs_param['off']) * (
                                                   (1 + cost_infl) ** (step_y * y_index))
                # Arc regulated tariff (USD/kcm)
                ### Note: the (ORD(y)-1) expression in GAMS is equivalent of (index(y)) in python, because python indexing starts from 0
                # "Arc expansion cost USD per kcm" inv_a(a,y)
                # arcs_param['inv_a'] = round((arcs_param['inv_a_base'] * arcs_param['i_cal'] * (1 - arcs_param['l_a']) * (arcs_param['len'] + max(0, (PipeOffshMult - 1)) * arcs_param['off']) * ((1 + cost_infl) ** (step_y * y_index)) / 365 / step_y),1)
                # the value of 'inv_a' is rounded by 1 decimal to be identical with GAMS input
                arcs_param['inv_a'] = (arcs_param['inv_a_base'] * arcs_param['i_cal'] * (1 - arcs_param['l_a']) * (
                            arcs_param['len'] + max(0, (PipeOffshMult - 1)) * arcs_param['off']) * (
                                                   (1 + cost_infl) ** (step_y * y_index))) / 365 / step_y

                if int(y) == 2015:
                    # "Arc expansion lim  BCM/yr" For First year, year = 2015
                    arcs_param['d_a_max1'] = arcs_param['d_max1'] / (
                                1 - arcs_param['l_a'])  # for the first year (Year = 2015)
                elif int(y) == 2020:
                    arcs_param['d_a_max2'] = arcs_param['d_max2'] / (1 - arcs_param['l_a'])  # for the second year
                elif int(y) > 2020:
                    arcs_param['d_a_max3'] = arcs_param['d_max3'] / (
                                1 - arcs_param['l_a'])  # for the third year and above
                else:
                    return 'notDefined'
                return arcs_param
                # Pass the value of year to compute parameters affected by year

            arc_param = arc_customize(last_year)

        except Exception as e:
            print(e)
            print("Check computations for Arc's parameters")
            ############ The set and parameter values for LNG #############################
            # Get the LNG Arc Names, to do that:
            # In Arcs dataframe, we take the first column (Arcs name) and check if the row contains 'Ship'. Then make a new dataframe with the name of LNG arc, change the header and reset the index
        try:
            LNG = Arcs[Arcs['0'].str.contains('Ship')].rename(columns={'0': 'LNG'}).reset_index(drop=True)
            # LNG Arc Parameters
            # Get the scalar parameter of LNG from 'Other' dataframe as new added columns to LNG dataframe
            LNG['cost_av_base'] = float(Others['BFShip'])  # are defined as integer in GAMS
            LNG['l_av_base'] = float(Others['BLShip'])
            LNG['l_av_base'] = LNG['l_av_base'].replace({'EPS': 0.0})  # Set EPS value to 0

            # Get the value for LNG distance parameter
            # Read data for LNG distance (Matrix from excel), the dataframe is LNGDistances, and it needs to clean and set the header
            LNGDistances_cleaned = clean_header(
                LNGDistances)  # this includes the liq node and reg node of two sides of LNG arc
            # Get the Nodes of two sides of LNG (Liq node and Reg node)
            # set the first column as index
            LNGDistances_cleaned = LNGDistances_cleaned.loc[:, 'None':]
            LNGDistances_cleaned.set_index('None', inplace=True)

            # Get the labels, columns and cell values and make dataframe
            a = np.repeat(LNGDistances_cleaned.index, len(LNGDistances_cleaned.columns))  # it takes rows labels (index)
            b = np.tile(LNGDistances_cleaned.columns, len(LNGDistances_cleaned))  # Takes columns name
            c = LNGDistances_cleaned.values.ravel().astype(
                float)  # it takes the value for cells (cross-product of row and columns) and convert to float

            # no:start node, ni:end node
            lng_data = {'no': a, 'ni': b,
                        'dist': c}  # Make a dictionary from index, columns and value (a,b,c) respectively

            # Note: 'no' as Start node and 'ni' as end node
            # finally, make a dataframe for LNG start and End node with distance Parameter
            df_lng = pd.DataFrame(lng_data,
                                  columns=['no', 'ni', 'dist'])  # make a dataframe with column name (nl, nr, dist)

            # Add other parameters of LNG
            # The length of LNG arc is equal to the distance parameter between two nodes (liq and Reg)
            df_lng['len'] = df_lng['dist']
            df_lng['c_cal'] = 1  # according to the GAMS code, the cost calibration parameter for LNG is 1

            # merge two dataframe (LNG) and (df_lng), to index parameter over LNG arc and related nodes
            # df_lng = df_lng.reset_index(drop=True)
            LNG_Arc_data = pd.concat([LNG, df_lng], axis=1)

            # add new parameter of LNG (l_a) for LNG
            LNG_Arc_data['l_av'] = round(LNG_Arc_data['l_av_base'] * LNG_Arc_data['len'], 3)

            # These parameters are not computed for LNG: ['inv_av'], ['cap_a'], ['d_a_max1'], ['d_a_max2'],['d_a_max3']

            def alert_LNG_y(y):
                y_index = get_indx(y_df, y)  # Get the index position of the Year
                # in the table of LNG data values, there is no value for LNG_Arc_data['off']. Here, I did not consider this value to avoid invalidation of the result.
                LNG_Arc_data['cost_av'] = LNG_Arc_data['cost_av_base'] * LNG_Arc_data['c_cal'] * (
                            1 - LNG_Arc_data['l_av']) * (LNG_Arc_data['len']) * ((1 + cost_infl) ** (step_y * y_index))
                ### NOTE: This term is omitted for LNG [max(0,dat_oth('PipeOffshMult')-1)*dat_a(a,n_o,n_i,"off")] to get the value of 'cost_av', because there is no value for 'off'
                return LNG_Arc_data

            LNG_Arc_data = alert_LNG_y(last_year)
            LNG_Arc_data = LNG_Arc_data.rename(
                columns={'LNG': 'Arc', 'no': 'Start', 'ni': 'End', 'l_av': 'l_a', 'cost_av': 'cost_a'})
            LNG_nod = LNG_Arc_data[['Arc', 'Start', 'End']]

        except Exception as e:
            print(e)
            print('Check codes for LNG data')
            # End of Arcs parameter value calculations
        #############################################################
        ############################## Calibration Data values ############################################
        # Read the GlobalLoss Rate (GlobalLoss.csv) saved as GlobalLoss dataframe
        try:
            Global_Loss = GlobalLoss.copy()
            Global_Loss = Global_Loss.rename(columns={'0': 'Year', '1': 'l_glob'})  # Change the header name
            Global_Loss['l_glob'] = Global_Loss['l_glob'].replace({'EPS': 0})  # Set EPS value to 0
            Global_Loss['Year'] = Global_Loss['Year'].astype(
                str)  # Change the datatype of year value for indexing in pyomo
            # convert dataframe to dictionary with (parameter, value) to get the loss rate for each year
            GlobalLoss_dict = Global_Loss.set_index('Year').T.to_dict(
                'series')  # Remove the index from dataframe, transpose and convert to dictionary as list format

            def get_golbloss(y):
                l_glob = int(GlobalLoss_dict[y])  # the globalLoss rate for 'y'
                return l_glob

            # l_glob(y) in GAMS
            l_glob = get_golbloss(
                last_year)  # pass the value of global parameter for year and get the loss rate for that specific year
        except Exception as e:
            print(e)
            print('Check the Global loss data values')
            # Price calibration data value is read from PriceCalib dataframe
            # Auxiliary set for price calibration
        ################################ Price Calibration #############################################
        try:
            PriceCalib_cleaned = clean_header(PriceCalib)  # Clean the dataframe header
            PriceCalib_cleaned = PriceCalib_cleaned.rename(columns={'NODE': 'N'})
            PriceCalib_cleaned['price'] = PriceCalib_cleaned['price'].replace('0', '150').astype(
                float)  # replace 0 value of price with 150 and change type to float

            def priceCal_df(y):
                PriceCalib_df = PriceCalib_cleaned.copy()  # make a copy of dataframe
                PriceCalib_df[y] = PriceCalib_df[y].replace('0', '1').astype(
                    float)  # get the value under the column of year multiplication and convert to float
                # calculate the price growth
                PriceCalib_df['pr_grow'] = PriceCalib_df[y] * (
                        (1 + pr_infl) ** (step_y * y_index))  # create new column for pr_grow and put the computed value
                # Calculate the reference price at nodes
                PriceCalib_df['ref_pr'] = PriceCalib_df['price'] * PriceCalib_df[
                    'pr_grow']  # create new column as ref_pr and put the computed value

                # change the name of column year (y) to prCal_y for later merging
                for col in PriceCalib_df.columns:
                    name = 'prCal_' + str(y)  # get the string of global parameter year and add to 'prCal_' for later
                    PriceCalib_cl = PriceCalib_df.rename(
                        columns={y: name})  # get the column with value of year and change to customized name
                return PriceCalib_cl

            PriceCalib_final = priceCal_df(int(last_year))

        except Exception as e:
            print(e)
            print('Check Price Calibration block')
            ### function for datatype change and replacing 0 with 1 for year(y)
        ##################################### ProdCapacity Calibration #################################

        try:
            # Production Capacity Calibration from ProdCapCalib dataframe
            ProdCapCalib_cleaned = clean_header(ProdCapCalib)
            ProdCapCalib_cleaned = ProdCapCalib_cleaned.rename(columns={'NODE': 'N'})

            # ProdCapCalib_cleaned = ProdCapCalib_cleaned.rename(columns={2015:'y2015_capCal', 'R1':'R1_sh'})
            ### function to change data type and replace 0 with 1 for the value of Resources (R) and Year (y)
            def prodcap_cal(y, r):
                prodcap_df = ProdCapCalib_cleaned.copy()
                prodcap_df[y] = prodcap_df[y].replace('0', '1').astype(float)  # multiplication factor of year capacity

                prodcap_df[r] = prodcap_df[r].astype(float)  # share of Resources

                # Rename the column of year (2015) to (prodCal_) for further merging
                for col in prodcap_df.columns:
                    name = 'prodCal_' + str(y)
                    ProdCapCal_df = prodcap_df.rename(columns={y: name})

                return ProdCapCal_df
                # Get the value of Resources dataframe to pass to the following function (for example getting (R1) from Resources
                # This is just tested on tests with one value for Rs (only one production resource)

            for R in df_r['R']:
                R == df_r['R']

            ProdCapCalib_final = prodcap_cal(int(last_year),
                                             R)  # includes production nodes, # we pass integer of y -> 2015 because the header is integer

        except Exception as e:
            print(e)
            print('Check the ProdCapacity Calibration block')

        ############################## ProductCost Calibration ########################
        try:
            # production cost Calibration from ProdCostCalib dataframe
            # Merge first two rows for names
            ProdCostCalib.columns = (ProdCostCalib.iloc[0] + '_' + ProdCostCalib.iloc[1])  # join the labels in two rows
            ProdCostCalib_cleaned = ProdCostCalib.iloc[2:].reset_index(drop=True)  # set the header
            ProdCostCalib_cleaned['base_cost'] = ProdCostCalib_cleaned['base_cost'].astype(float)
            ProdCostCalib_cleaned = ProdCostCalib_cleaned.rename(columns={'None_NODE': 'N'})

            ### function for updating ProdCostCalib dataframe (datatype and replacing 0 with 1) for Year(R) and Resource (R)
            def prodcost_cal(y, r):
                prodcost_df = ProdCostCalib_cleaned.copy()
                for col in prodcost_df.columns:
                    if y in col:  # it will get all columns that have 'y' in their header
                        prodcost_df[col] = prodcost_df[col].replace('0', '1').astype(
                            float)  # multiplication factor of year cost calibration
                    elif r in col:  # it will get columns for resources
                        if 'c' in col:  # columns for r_c (linear cost)
                            prodcost_df[col] = prodcost_df[col].astype(float)  # linear increase
                        elif 'q' in col:  # columns for r_q (quadratic cost)
                            prodcost_df[col] = prodcost_df[col].replace('EPS', 0).astype(float)  # quadratic increase
                return prodcost_df

            ## Note: notice to the headers, since this function, gets the headers including (y) for years value (2015,2020,...)
            ProdCostCalib_final = prodcost_cal(last_year, R)

        except Exception as e:
            print(e)
            print('Check The ProductCost Calibration block ')
        ####################### Projection data ####################
        # Read Projection data values for WEO from WEO_ROW_scenarios dataframe
        # Clean the header and set the column names
        # Merge first two rows for names
        try:
            WEO_ROW_scenarios.columns = (WEO_ROW_scenarios.iloc[0] + '_' + WEO_ROW_scenarios.iloc[1])
            WEO_ROW = WEO_ROW_scenarios.iloc[2:].reset_index(drop=True)
            WEO_ROW = WEO_ROW.rename(
                columns={WEO_ROW.columns[0]: 'Scen', WEO_ROW.columns[1]: 'Rgn', WEO_ROW.columns[2]: 'N',
                         WEO_ROW.columns[3]: 'CN'})

            # Read Projection data for SETNav
            # Clean the header and set the column names
            # Merge first two rows for names
            SETNav_EU_scenarios.columns = (SETNav_EU_scenarios.iloc[0] + '_' + SETNav_EU_scenarios.iloc[1])
            SETNav_EU = SETNav_EU_scenarios.iloc[2:].reset_index(drop=True)
            SETNav_EU = SETNav_EU.rename(
                columns={SETNav_EU.columns[0]: 'Scen', SETNav_EU.columns[1]: 'Rgn', SETNav_EU.columns[2]: 'N',
                         SETNav_EU.columns[3]: 'CN'})

            weo_setnav = pd.concat([WEO_ROW, SETNav_EU], ignore_index=True)  # concat two dataframe for scenarios,
            weo_Set_df = weo_setnav[weo_setnav['Scen'].isin(
                [weo, setnav])]  # Get the rows related to defined scenario in 'Setting.txt' file

            ## General function to change the data type for Projection data values, prod, cons, elast
            def scenario_dtype_convert(df):
                for col in df.columns:
                    if 'Pro' in col:
                        df[col] = df[col].astype(float)
                    elif 'Cons' in col:
                        df[col] = df[col].astype(float)
                    elif 'Elast' in col:
                        df[col] = df[col].replace(0, -0.4).astype(float)
                return df

                # NOTE: The name in headers of Excel tables are important
                # Convert data type to float

            weo_Set_df = scenario_dtype_convert(weo_Set_df)

            # Production Growth
            p_grow = (1 + cost_infl) ** (step_y * y_index)

        except Exception as e:
            print(e)
            print('Check the Projection data block')
        # Note: the headers are changed based on the ordered index of excel table
        ################################################################
        ################# Consumption parameters ####################
        # merge four dataframe on Nodes, since other three dataframe (priceCalib, prodCostCalib and ProdCapCalib include produciton nodes, the final dataframe also includes production Nodes)
        try:
            dfs_cons = [weo_Set_df,
                        PriceCalib_final]  # PriceCalib_final includes ref_price needed for computing consumption paramters
            df_cons = reduce(lambda left, right: pd.merge(left, right, on='N'),
                             dfs_cons)  # merge four dataframe on Nodes, since other three dataframe (priceCalib, prodCostCalib and ProdCapCalib include produciton nodes, the final dataframe also includes production Nodes)

            ################### The glob reference of consumption and production ######################
            def ref_prod_glob(y):

                # The reference global production is the sum of reference production over producer nodes
                glob_prod = round(weo_Set_df['Prod_' + str(y)].sum())

                return glob_prod

            ref_p_glob = ref_prod_glob(int(last_year))

            def ref_cons_glob(y):

                # The reference global consumption is the sum of reference consumption over consumer nodes
                glob_cons = round(weo_Set_df['Cons_' + str(y)].sum())

                return glob_cons

            ref_c_glob = ref_cons_glob(int(last_year))

            # compute c_adj for ROW and ROE and EU
            def c_adj_func(df):
                c_adj = []  # an empty list for value assignment
                c_adj_mismatch = ref_p_glob / ref_c_glob

                for row in df['Rgn']:
                    if row == 'ROE':
                        c_adj.append(1)
                    elif row == 'EU':
                        c_adj.append(1)
                    elif row == 'ROW':
                        c_adj.append(c_adj_mismatch)
                df['c_adj'] = c_adj
                return df

            cons_df = c_adj_func(df_cons)
        except Exception as e:
            print(e)
            print('Check the Consumption parameter')

        ################# Production parameters ####################
        # merge four dataframe on Nodes, since other three dataframe (priceCalib, prodCostCalib and ProdCapCalib include produciton nodes, the final dataframe also includes production Nodes)
        try:
            dfs_prod = [weo_Set_df, PriceCalib_final, ProdCapCalib_final, ProdCostCalib_final]
            df_prod = reduce(lambda left, right: pd.merge(left, right, on='N'), dfs_prod)

            ## compute the production parameters (production capacity, linear cost of production, quadratic production cost)
            # This function gets the dataframe name, value of year (in string format) and resource (string format)
            def prod_param(df, y, r):
                # GAMS code: cap_p(n_p,r,y) = ref_p(n_p,y)*cap_p_cal(n_p,r)*cap_p_cal(n_p,y); For Nodes with production of its share of resources and multiply by the year calibration multiplication factor
                # Production Cap (cap_p)
                df['cap_p'] = df['Prod_' + str(y)] * df[r] * df['prodCal_' + str(y)]

                # GAMS code: cost_pq(n_p,r,y)$cap_p(n_p,r,y) = cost_p_cal(n_p,'base','cost')* cost_p_cal(n_p,r,'q')*cost_p_cal(n_p,'y',y)* p_grow(n_p,y)/cap_p(n_p,r,y);
                df['cost_pq'] = df['base_cost'] * df[str(r) + '_q'] * df['y_' + str(y)] * p_grow / df['cap_p']

                # GAMS code: cost_pl(n_p,r,y)$cap_p(n_p,r,y) = cost_p_cal(n_p,'base','cost')* cost_p_cal(n_p,r,'c')*cost_p_cal(n_p,'y',y)* p_grow(n_p,y);
                df['cost_pl'] = df['base_cost'] * df[str(r) + '_c'] * df['y_' + str(y)] * p_grow

                return df

            prod_df = prod_param(df_prod, last_year, R)

        except Exception as e:
            print(e)
            print('Check the Production parameters block')
        ##################### The reference consumption for ROW and ROE and EU
        try:
            def reference_consumption(df, y):
                # l_glob(y)
                l_glob = get_golbloss(str(y))  # to make sure it passes string format of y to get global loss (str())
                for row in df['Rgn']:
                    if row == 'ROE':
                        df['ref_c'] = df['Cons_' + str(y)] * df['c_adj']
                    elif row == 'EU':
                        df['ref_c'] = df['Cons_' + str(y)] * df['c_adj']
                    elif row == 'ROW':
                        df['ref_c'] = (1 - l_glob) * df['Cons_' + str(y)] * df['c_adj']
                return df

            cons_df = reference_consumption(df_cons, int(last_year))  # dataframe of reference consumption for global parameter year
        except Exception as e:
            print(e)
            print('Check the reference consumption for ROW and ROE and EU block')
        ################# The slope and intercept of demand curve ##################3
        try:
            def int_slp(df, y):
                df['slp'] = (-df['ref_pr']) / df['ref_c'] / df['Elast_' + str(y)]

                df['slp'] = df['slp'].replace(np.inf, 0.0).astype(float) # it replaces the infinite values in Slope with 0 then can compute the intercept value
                df['int'] = df['ref_pr'] + df['slp'] * df['ref_c']
                return df

            cons_df = int_slp(df_cons, int(last_year))
            print(cons_df)

        except Exception as e:
            print(e)
            print('Check The slope and intercept of demand curve block')
        ################### Market Power ######################
        try:
            MarketPower_df = clean_header(MarketPower)
            MarketPower_df = MarketPower_df.rename(columns={'None': 'CN'})

            mp_cn = MarketPower_df[['CN']].reset_index(drop=True)  # dataframe of countries with Market Power
            # convert data type
            MarketPower_df[['domestic', 'export', 'factor', 'ratio']] = MarketPower_df[
                ['domestic', 'export', 'factor', 'ratio']].replace('EPS', 0).astype(float)

        except Exception as e:
            print(e)
            print('Check the Market Power block')

        ####################### the mapping of countries with nodes type #############################

        try:
            # supplier has access to its own np
            s_np = NodesData_cleaned.loc[NodesData_cleaned['P'] == '1', ['CN', 'N']].reset_index(
                drop=True)  # dataframe of suppliers with np
            # mp_cn
            ## suppliers has access to own production node, merge dataframe of supplier_prod_node (s_np) with countries from MarketPower table(mp_cn)
            t_acc_np = mp_cn.merge(s_np, on=[
                'CN'])  ## select common countries from country with prod_node and market countries
            t_acc_np = t_acc_np.rename(columns={'N': 'np'})

            ## supplier has access to consumption node
            nc = NodesData_cleaned.loc[NodesData_cleaned[
                                           'C'] == '1', 'N']  # List of Consumption Nodes, (nodes with consumption columns values (C) == 1)
            suppliers = NodesData_cleaned.loc[
                NodesData_cleaned['P'] == '1', 'CN']  # List of supplier Nodes, (nodes with production columns == 1)
            s_nc = [(i, j) for i in suppliers for j in nc]  # creates list of tuples from product of two lists
            s_acc_nc = pd.DataFrame(s_nc, columns=['CN', 'N'])  # supplier has access to all consumption nodes
            t_acc_nc = s_acc_nc.merge(t_acc_np, on=['CN'])
            t_acc_nc = t_acc_nc.rename(columns={'N': 'nc'})  # supplier access to prod_node and cons_node

            # CN_ConsN = [(i, j) for i in suppliers for j in n_cons]
            # CN_ConsN = pd.DataFrame(CN_ConsN, columns=['CN', 'N'])  # dataframe of Countries accessed to consumption nodes

            # supplier has access to regular pipeline
            pipe_arc = ArcsData_cleaned.loc[ArcsData_cleaned['P'] == '1', 'Arc']
            s_ap = [(x, y) for x in suppliers for y in pipe_arc]
            s_ap_df = pd.DataFrame(s_ap, columns=['CN', 'ap'])
            t_acc_ap = s_ap_df.merge(t_acc_nc, on=['CN'])  # map supplier with market power with pipeline arc
            # supplier has access to pipeline Arcs

            # supplier has access to liq node and liq arc,
            # if the arc is connected to one of supplier's np --> the start node of arc == np of supplier
            # map of liq arc with start and end node
            map_al_no_ni = arc_node('Liq').reset_index(drop=True)

            # supplier has access to regas and cons nodes, and to shipping and regas arcs.
            # the access to regas node and connected cons_node, if supplier access to liq arc and cons_node is an own np not excluded
            map_ar_no_ni = arc_node('Regas').reset_index(drop=True)

            t_al_ln = map_al_no_ni.merge(t_acc_ap, how='cross')
            # t_al_ln = t_al_ln.rename(columns={'Arc':'ar'})
            t_ar_rn = map_ar_no_ni.merge(t_acc_ap, how='cross')

            supplier_aral = pd.concat([t_al_ln, t_ar_rn], ignore_index=True)
            # join the dataframes of liq arc and reg arc which are merged with supplier

            suppler_acc_liq = supplier_aral[supplier_aral['Start'] == supplier_aral['CN']].reset_index(
                drop=True)  # supplier has access to liq arc and liq node
            suppler_acc_reg = supplier_aral[supplier_aral['End'] == supplier_aral['nc']].reset_index(
                drop=True)  # supplier has access to reg arc and reg node, if the end node of reg arc is same as cons_node accessed by supplier
            supplier_acc_liq_reg = pd.concat([suppler_acc_liq, suppler_acc_reg],
                                             ignore_index=True)  # concat two dataframes, supplier with access to liq and supplier with access to Regas
            supplier_acc_liqreg = supplier_acc_liq_reg.groupby('CN').filter(lambda g: len(g) > 1).drop_duplicates(
                subset=['Arc', 'Start', 'End', 'CN', 'ap', 'nc', 'np'],
                keep="first")  # select supplier that has access to both liq and reg arcs and nodes
            # grouped by 'CN' and filter by common values in column 'CN'

            supplier_liqNode = supplier_acc_liqreg[['CN', 'End']].groupby('CN').filter(
                lambda g: len(g) != 1).drop_duplicates(subset=['CN', 'End'])
            supplier_liqNode = supplier_liqNode.rename(columns={'End': 'N'})
            supplier_regNode = supplier_acc_liqreg[['CN', 'Start']].groupby('CN').filter(
                lambda g: len(g) != 1).drop_duplicates(subset=['CN', 'Start'])
            supplier_regNode = supplier_regNode.rename(columns={'Start': 'N'})
            supplier_LiqRegNode = supplier_regNode.append(supplier_liqNode, ignore_index=True)

            t_acc_n = pd.concat([s_acc_nc, supplier_LiqRegNode], ignore_index=True)
            t_acc_n = t_acc_n.drop_duplicates(subset=['CN', 'N'],
                                              keep="first")  # It drops duplicate rows in ['CN' , 'N']

        except Exception as e:
            print(e)
            print('Check the mapping of countries with nodes type')

            ######################  Cour value  ########################
            # determine export or domestic of market power exertion
        try:
            t_acc_nc['cour_type'] = np.where((t_acc_nc['nc'] == t_acc_nc['np']), 'domestic', 'export')

            # join MarketPower_df with t_acc_nc on 'CN'
            MP_t_join = t_acc_nc.merge(MarketPower_df, on=['CN'])

            def get_cour(df, y):

                # Market power moderation in later years
                # cour(t,n_c,y)=max(cour(t,n_c,y)*dat_mp(t,'ratio'),cour(t,n_c,y)*power(dat_mp(t,'factor'),ORD(y)-1));
                # first creat new column from multiplication of ratio on domestic and export value, same for multiplication of factor on thoese
                df['dom_ratio'] = df['domestic'] * df['ratio']
                df['dom_factor'] = df['domestic'] * (df['factor'] ** (step_y * y_index))
                df['exp_ratio'] = df['export'] * df['ratio']
                df['exp_factor'] = df['export'] * (df['factor'] ** (step_y * y_index))

                # get the max value between multiplication values based on condition (domestic or export)
                df['cour'] = np.where((df['cour_type'] == 'domestic'), df[['dom_ratio', 'dom_factor']].max(axis=1),
                                      df[['exp_ratio', 'exp_factor']].max(axis=1))
                # convert data to float and round by 2 decimal
                df['cour'] = df['cour'].astype(float).round(2)
                return df

            mp_t = get_cour(MP_t_join, int(last_year))

        except Exception as e:
            print(e)
            print('Check the Cour value')

        ############################ Save the second 'CSV' files for GGM_pyomo model #################################
        # Build dataframe with the value of parameters and the relevant indices.
        # Then save dataframe in 'CSV' format inside different folder 'ggm'
        ################ merge year and resource dataframe #####################
        try:
            # def set_index(y,r):
            def get_y_r(y, r):
                for row in y_df['Year']:
                    if row == y:  # Get rows with values of Year (y)
                        df_year = y_df['Year']
                        # print(type(df_y))
                    else:
                        pass
                for row in df_r['R']:
                    if row == r:  # Get rows with values of Resource (R)
                        df_resource = df_r['R']
                        # print(type(df_r))
                    else:
                        pass
                df = pd.concat([df_year, df_resource], axis=1)  # Merge to series and convert to dataframe
                return df

            # Pass specific value for year and resource in string format
            df_year_resource = get_y_r(last_year, R)
        except Exception as e:
            print(e)
            print('Check merge year and resource dataframe block')

        ###################### Combine parameter values with year and R ######################################
        try:
            def combine_y_r(df):
                combined_df = pd.concat([df, df_year_resource], axis=1)
                combined_df['Year'] = combined_df['Year'].fillna(method='ffill')
                combined_df['R'] = combined_df['R'].fillna(method='ffill')
                return combined_df

            # Constant production cost, for production node
            cost_pl = combine_y_r(prod_df)
            cost_pl = cost_pl[['N', 'R', 'Year',
                               'cost_pl']]  # get relevant columns in order of indices and then the value of parameter

            # Production capacity, for production node
            cap_p = combine_y_r(prod_df)
            cap_p = cap_p[['N', 'R', 'Year', 'cap_p']]

            # Linear production cost, for production node
            cost_pq = combine_y_r(prod_df)
            cost_pq = cost_pq[['N', 'R', 'Year', 'cost_pq']]

            # arc usage cost
            ArcsData_yr = combine_y_r(arc_param)
            cost_arcs = ArcsData_yr[['Arc', 'Year', 'cost_a']].dropna().reset_index(drop=True)
            av_yr = combine_y_r(LNG_Arc_data)
            av_cost = av_yr[['Arc', 'Year', 'cost_a']]  # .rename(columns={'LNG':'Arc', 'cost_av':'cost_a'})
            cost_a = pd.concat([cost_arcs, av_cost], ignore_index=True)

            # the capacity of Arc
            cap_a = ArcsData_yr[['Arc', 'Year', 'cap_a']].dropna().reset_index(drop=True)

            # d_a_max
            for i in ArcsData_yr['Year']:
                if int(i) == 2015:
                    # "Arc expansion lim  BCM/yr" For First year, year = 2015
                    d_a_max1 = ArcsData_yr[['Arc', 'Year', 'd_a_max1']].dropna().reset_index(
                        drop=True)  # for the first year (Year = 2015)
                elif int(i) == 2020:
                    d_a_max2 = ArcsData_yr[['Arc', 'Year', 'd_a_max2']].dropna().reset_index(
                        drop=True)  # for the second year
                elif int(i) > 2020:
                    d_a_max3 = ArcsData_yr[['Arc', 'Year', 'd_a_max3']].dropna().reset_index(
                        drop=True)  # for the third year and above
                else:
                    return 'notDefined'

            # Arc Loss rate
            l_arcs = ArcsData_yr[['Arc', 'Year', 'l_a']].dropna().reset_index(drop=True)
            av_l = av_yr[['Arc', 'Year', 'l_a']]
            l_a = pd.concat([l_arcs, av_l], ignore_index=True)

            # arc investment cost
            inv_a = ArcsData_yr[['Arc', 'Year', 'inv_a']].dropna().reset_index(drop=True)

            # cour/ market power
            mp_t_df = combine_y_r(mp_t)
            cour = mp_t_df[['np', 'nc', 'Year', 'cour']]
            # cour_mera = cour_value.iloc[0:8] # first five rows of dataframe

            # slope
            cons_df_y = combine_y_r(cons_df)
            slp = cons_df_y[['N', 'Year', 'slp']]

            # intercept
            intercept = cons_df_y[['N', 'Year', 'int']]

            # Discount rate for y
            # update df with year and resource, add new column for discount rate for year
            disc = df_year_resource.copy()
            # disc(y) = 1 / power(1 + disc_rate, step_y * (ORD(y) - 1));
            disc['disc'] = 1 / (1 + disc_rate) ** (step_y * y_index)
            disc_y = disc[['Year', 'disc']]

        except Exception as e:
            print(e)
            print('Check the Combine parameter values with year and R')

        #######################
        try:
            # indicator for arcs with start and end node
            arc_start = ArcsData_cleaned[['Arc', 'Start']]
            arc_end = ArcsData_cleaned[['Arc', 'End']]
            av_ni = LNG_Arc_data[['Arc', 'End']]  # .rename(columns={'LNG':'Arc','ni':'End'}) # end
            av_no = LNG_Arc_data[['Arc', 'Start']]  # .rename(columns={'LNG':'Arc','no':'Start'}) # start

            a_e = pd.concat([arc_end, av_ni], ignore_index=True)
            a_s = pd.concat([arc_start, av_no], ignore_index=True)

            Arcs_prl = ArcsData_cleaned['Arc'] # List of All Regas, Pipe and Liq arcs
            ap = ArcsData_cleaned.loc[ArcsData_cleaned['P'] == '1', ['Arc','Start']]  # List of Consumption Nodes, (nodes with consumption columns values (C) == 1)
            ar = ArcsData_cleaned.loc[ArcsData_cleaned['R'] == '1', ['Arc','Start']]  # List of Production Nodes, (nodes with production columns (P) == 1)
            al = ArcsData_cleaned.loc[ArcsData_cleaned['L'] == '1', ['Arc','Start']]  # List of liquefaction Nodes, (nodes with L columns (L) == 1)
            av = LNG_Arc_data[['Arc', 'Start']]  # List of regasifaction Nodes, (nodes with R columns (R) == 1)

        except Exception as e:
            print(e)
            print("Check Arc's indicators")

            ################## Save for the model ########################
            # set the working directory to new sub_folder for saving parameters with indices
        try:
            ggm_path = os.path.join(test_dir, 'ggm')
            if not os.path.exists(ggm_path):
                os.makedirs(os.path.join(test_dir, ggm_path))

            # name = [x for x in globals() if globals()[x] is arc_start][1]  # get name of dataframe, [0] will generate ('_') but changing ro [1] will get the name
            # arc_start.to_csv(f'{ggm_path}/{name}.csv', index=False)  # save dataframe to 'csv' format with the name of dataframe
            # To make it simple, there is a list of the name of sets, indicator and parameters that are needed by the pyomo. Then each dataframe is saved with related name as 'CSV*
            new_dict = {'Nodes': Nodes, 'nc': n_cons, 'np': n_prod, 'nl': n_liq, 'nr': n_reg, 'cn': countries,
                        's': suppliers,
                        'map_n_cn': map_n_cn, 'map_cn_rgn': map_cn_rgn, 'map_n_rgn': map_n_rgn,
                        'map_n_cn_rgn': map_n_cn_rgn,
                        'pipe_nod': pipe_nod, 'liq_nod': liq_nod, 'reg_nod': reg_nod, 'LNG_nod': LNG_nod, 'Arcs_prl':Arcs_prl,
                        'Arcs': Arcs, 'ap': ap, 'al': al, 'ar': ar, 'av': av, 'a_s': a_s, 'a_e': a_e,
                        't_acc_np': t_acc_np, 't_acc_nc': t_acc_nc, 't_acc_ap': t_acc_ap, 't_al_ln': t_al_ln,
                        't_ar_rn': t_ar_rn, 't_acc_n': t_acc_n,
                        'cost_pl': cost_pl, 'cap_p': cap_p, 'cost_pq': cost_pq, 'cost_a': cost_a, 'cap_a': cap_a,
                        'l_a': l_a, 'inv_a': inv_a, 'cour': cour, 'slp': slp, 'int': intercept,
                        'disc': disc_y,'Years' : Years, 'Resources': Resources}
            for key, val in new_dict.items():
                df = pd.DataFrame(new_dict[key])  # make dataframe from groups (dictionary) based on its keys
                df.to_csv(f'{ggm_path}/{key}.csv', index=False)  # header=0)

            if int(last_year) == 2015:
                d_a_max1.to_csv(f'{ggm_path}/d_a_max1.csv', index=False)  # header=0)
            if int(last_year) == 2020:
                d_a_max2.to_csv(f'{ggm_path}/d_a_max2.csv', index=False)  # header=0)
            if int(last_year) > 2020:
                d_a_max3.to_csv(f'{ggm_path}/d_a_max3.csv', index=False)  # header=0)

        except Exception as e:
            print(e)
            print('Check the save for the model block')
        return file
    #print("The parameters are successfully computed and stored as 'CSV' format")

except Exception as e:
    print(e)
    print('Check the nested try-catch blocks')

clean_data(file)
print("The end of reading and cleaning data")

### For prodcap_cal and prodcost_cal, it summarizes the value of R1,R2,R3
### For Production Cap (cap_p), in prod_df
### WEO and SETNav scenarios, have single data_proj.xlsx in dataset Test folders.
### For general model, there should be condition for the value of 'weo' and 'setnav' which are read from 'Setting.txt'
########### Make a dictionary from all dataframes ############
# cost_pl,cap_p,cost_pq,cost_arc,cap_a,d_a_max1,arc_loss,inv_a,cour_value,slope,intercept,disc_y
# result = {"df{}".format(i): dict(df.values.tolist()) for i, df in [cost_pl,cap_p,cost_pq,cost_arc,cap_a,d_a_max1,arc_loss,inv_a,cour_value,slope,intercept,disc_y]}

# To get the index of year, the value of global parameter in 'Setting.txt' files should be defined in range in the excel file. Otherwise it returns the index error
